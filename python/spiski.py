import openpyxl
from openpyxl.styles import Font,PatternFill
errors=0
data2=[]

# цвета
red="FF0000"

# номера ячеек
acheykis = {
    'fimale': 'L',
    'name': 'M',
    'fio': 'N',
    'date_of_birth': 'O',
    'pasport_n': 'P',
    'gragd': 'Q',
    'date_of_end': 'R',
    'pol': 'S'
}
class Ecxel():
    # подсчитываем все листы
    def lists(self,list_number):
        list0, list2=list_number,list_number
        kolvo=0
        if list_number=='all':
            list2=len(kolvo_lists)+1
            list0=1
            kolvo = 1
        for b in list_number:
            if b == '-':
                list_number=list_number.split('-')
                list0=list_number[0]
                list2=list_number[1]
                kolvo=1

        if kolvo==1:
            list0, list2 = int(list0)-1, int(list2)-1
        else:
            list0, list2 = int(list0)-1, int(list2)

        all_lists = list2 - list0
        return list0, list2, all_lists

    # определение пола
    def pol_cheker(self,name):
        name1 = name
        if len(name1)>=1:
            if name1[-1] == 'A':
                pol = 'MRS'
                return pol
            else:
                pol = 'MR'
        else:
            pol='None'
        return pol

    # убираем лишнее из гражданства
    def gragd_deleter(self, gragd):
        for bukva in gragd:
            if bukva.isdigit() == False:
                gragd = gragd[1:]
            else:
                break
        return gragd

    # проверяем гражданство
    def gragdanstvo_cheker(self, gragd):
        gragd = str(gragd)
        gragd = gragd.lower()
        if gragd[:2] == 'np':
            gragd = gragd[2:]
        fistn = gragd[0]
        if fistn == '4':
            print(1)
            gragd1 = 'TJK'
        elif fistn == '5' or fistn == '6' or fistn == '7':
            print(2)
            gragd1 = 'RUS'
        elif fistn == 'a' or fistn == 'f' or fistn == 'k' or fistn == 'а':
            print(3)
            gragd1 = 'UZB'
        else:
            gragd1 = 'None'
        return gragd1

    # определяем гражданство
    def finel_gragd(self,gragd):
        gragd1=Ecxel.gragdanstvo_cheker(self,gragd)
        if gragd1=='None':
            gragd = Ecxel.gragd_deleter(self, gragd)
            gragd1 = Ecxel.gragdanstvo_cheker(self, gragd)
        return gragd1

    # удаляем ненужный мусор
    def deleter(self, data):
        # олучаем полную строчку
        data = data.split(' ')
        data2 = []
        for i in data:
            i = i.split('/')
            for g in i:
                data2.append(g)
        data = data2
        for i in range(len(data)):
            i = 0
            for slovo in data:
                if len(slovo) <= 3:
                    data.pop(i)
                    break
                if slovo == 'TJ' or slovo == 'RUS' or slovo == 'UZB' or slovo == 'UZ':
                    data.pop(i)
                    break
                i = i + 1
        g = 0
        for bukva in data[1]:
            if bukva.isdigit() == True:
                if len(data) >= 4:
                    buf = data[1]
                    fimale = buf[:g]
                    dada_of_b = buf[g:]
                    data.append('')
                    data[4] = data[3]
                    data[3] = data[2]
                    data[2] = dada_of_b
                    data[1] = fimale
                    break
            g = g + 1
        return data

    # ерекрашиваем клетки и подсчитываем ошибки
    def paiter(self,color,number,acheka):
        global errors
        if color=='red':
            list1[acheykis['fimale'] + str(acheka)].fill = PatternFill(start_color=red, end_color=red,
                                                   fill_type="solid")
            list1[str(number) + str(acheka)].fill = PatternFill(start_color=red, end_color=red,
                                                        fill_type="solid")
        errors=errors+1
        return errors

    # проверяем и исправляем даты
    def time_cheker(self, time, number):
        time = str(time)
        if time[-1] == '.':
            time = time[:-1]
        if len(time) > 10:
            time = time[:10]
        buf = time.split('.')
        if len(buf) < 3:
            buf2 = ''
            for i in buf:
                buf2 = buf2 + i
            day = buf2[:2]
            month = buf2[2:4]
            year = buf2[4:]
            time = day + '.' + month + '.' + year
        if len(time) == 8:
            if number == 0:
                if int(time[-2:]) <= 30:
                    year = '20' + time[-2:]
                else:
                    year = '19' + time[-2:]
                time = time[:-2]
                time = time + year
            if number == 1:
                year = '20' + time[-2:]
                time = time[:-2]
                time = time + year
        return time

    # разделение по столбикам
    def colomn_seperation(self,data,acheka,acheka_otchet):
        global errors
        acheka_otchet.append(acheka)
        if data!=None:
            if len(data)>=1:
                data=Ecxel.deleter(self,data)
                print(data)
                acheka_otchet.append(data)
                name=data[0]
                list1[acheykis['fimale']+str(acheka)].value=name

                # ставим имя фамилию и пол
                if len(data)>=2:
                    fimale=data[1]
                    list1[acheykis['name']+str(acheka)].value=fimale
                    Fio = str(name) + ' '+str(fimale)
                    list1[acheykis['fio'] + str(acheka)].value = Fio
                    pol = Ecxel.pol_cheker(self,fimale)
                    if pol=='None':
                        errors = Ecxel.paiter(self, 'red', acheykis['pol'], acheka)
                        acheka_otchet.append('пол,')
                    list1[acheykis['pol'] + str(acheka)].value = pol

                    # роставляем дату рождения
                    if len(data) >= 3:
                        date_of_birth=data[2]
                        date_of_birth=Ecxel.time_cheker(self,date_of_birth,0)
                        list1[acheykis['date_of_birth']+str(acheka)].value = date_of_birth
                        if len(date_of_birth)<10:
                            errors=Ecxel.paiter(self,'red',acheykis['date_of_birth'],acheka)
                            acheka_otchet.append('дата рождения,')

                        # роставляем паспортные данные и гражданство
                        if len(data) >= 4:
                            pasport_n=data[3]
                            if pasport_n[-1] == '.':
                                pasport_n = pasport_n[:-1]
                            list1[acheykis['pasport_n']+str(acheka)].value = pasport_n
                            if len(pasport_n)<6:
                                errors=Ecxel.paiter(self,'red',acheykis['pasport_n'],acheka)
                                acheka_otchet.append('паспортные данные,')
                            gragd=Ecxel.finel_gragd(self,pasport_n)
                            list1[acheykis['gragd']+str(acheka)].value = gragd
                            if gragd=='None':
                                errors=Ecxel.paiter(self,'red',acheykis['gragd'],acheka)
                                acheka_otchet.append('гражданство,')

                            # роставляем дату окончания паспотра
                            if len(data) >= 5:
                                data_end=data[4]
                                data_end = Ecxel.time_cheker(self, data_end,1)
                                list1[acheykis['date_of_end']+str(acheka)].value = data_end
                                if len(data_end) < 10:
                                    errors=Ecxel.paiter(self,'red',acheykis['date_of_end'],acheka)
                                    acheka_otchet.append('дата окончания,')
        return errors,acheka_otchet

    # расставляем загаловки
    def zagalovki(self):
        list1[str(acheykis['fimale'])+'2'].value = 'Фамилия'
        list1[str(acheykis['name'])+'2'].value = 'Имя'
        list1[str(acheykis['fio'])+'2'].value = 'Фамилия и имя'
        list1[str(acheykis['date_of_birth'])+'2'].value = 'Дата рожления'
        list1[str(acheykis['pasport_n'])+'2'].value = 'Номер паспотра'
        list1[str(acheykis['gragd'])+'2'].value = 'Гражданство'
        list1[str(acheykis['date_of_end'])+'2'].value = 'Срок действия'
        list1[str(acheykis['pol'])+'2'].value = 'Пол'

    def main(self,file_name,list_number,directory,exel_name):
        global list1,kolvo_lists,errors
        # открываем и сохдаем нужные переменнные
        otchet_of_work=open(str(directory)+'otchet_of_work '+exel_name+'.txt','w')
        otchet_name=str(directory)+'otchet_of_work '+exel_name+'.txt'
        errors_otchet=[]
        data_otchet = []
        errors,errors2=0,0
        str_otchet,str_errors_otchet='','Отчет об ошибках: \n'
        # название конечного файла
        save_name=file_name+' отредактированный'
        file = openpyxl.reader.excel.load_workbook(filename=file_name + '.xlsx')
        kolvo_lists=file.sheetnames
        # получаем первый последний и все листы
        list0,list2,all_lists=Ecxel.lists(self,list_number)
        # цикл листов в файле
        for l in range(1,all_lists+1):
            file.active = list0
            list1 = file.active
            Ecxel.zagalovki(self)
            i=2
            # цикл строчек на листе
            while True:
                i = i+1
                data=list1['B'+str(i)].value
                acheka_otchet = []
                acheka_otchet.append(l)
                errors2=errors
                errors,acheka_otchet=Ecxel.colomn_seperation(self,data,i,acheka_otchet)
                str_otchet=''
                for g in acheka_otchet:
                    g=str(g)
                    str_otchet=str_otchet+g+' '
                if errors2!=errors:
                    errors_otchet.append(str_otchet+'\n')
                data_otchet.append(str_otchet)
                otchet_of_work.write(str_otchet)
                otchet_of_work.write('\n')
                if data==None:
                    break
            list0 = list0 + 1
            otchet_of_work.write('\n')
        # сохраняем и закрываем все файлы
        for f in errors_otchet:
            str_errors_otchet=str_errors_otchet+str(f)
        otchet_of_work.write(str_errors_otchet)
        save_name=save_name+ '.xlsx'
        file.save(save_name)
        otchet_of_work.close()
        file.close()
        return save_name,all_lists,errors,otchet_name
